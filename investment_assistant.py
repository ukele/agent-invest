import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os


class InvestmentAssistant:
    def __init__(self):
        self.acoes = {
            'PETR4.SA': 'Petrobras',
            'VALE3.SA': 'Vale',
            'KLBN4.SA': 'Klabin',
            'BBAS3.SA': 'Banco do Brasil',
            'ITUB3.SA': 'Itau'
        }
        self.dados = []

    def buscar_dados_acao(self, ticker):
        """Busca dados da ação usando yfinance"""
        print(f"Buscando dados de {ticker}...")

        try:
            acao = yf.Ticker(ticker)

            # Informações atuais
            info = acao.info
            preco_atual = info.get('currentPrice') or info.get('regularMarketPrice')

            # Histórico de 52 semanas
            data_52_semanas_atras = datetime.now() - timedelta(days=365)
            historico = acao.history(start=data_52_semanas_atras, end=datetime.now())

            min_52_semanas = historico['Close'].min()
            max_52_semanas = historico['Close'].max()

            # Preco 12 meses atras
            preco_12m_atras = historico['Close'].iloc[0]
            valorizacao_12m = ((preco_atual - preco_12m_atras) / preco_12m_atras * 100) if preco_12m_atras > 0 else 0

            # Dividend Yield
            dividend_yield = info.get('dividendYield', 0)
            if dividend_yield:
                dividend_yield = dividend_yield * 100

            return {
                'Ticker': ticker,
                'Empresa': self.acoes.get(ticker, ticker),
                'Valor Atual (R$)': round(preco_atual, 2) if preco_atual else 'N/A',
                'Dividend Yield (%)': round(dividend_yield, 2) if dividend_yield else 'N/A',
                'Mín. 52 Semanas (R$)': round(min_52_semanas, 2),
                'Máx. 52 Semanas (R$)': round(max_52_semanas, 2),
                'Valorização 12m (%)': round(valorizacao_12m, 2)
            }

        except Exception as e:
            print(f"Erro ao buscar {ticker}: {str(e)}")
            return None

    def coletar_dados(self):
        """Coleta dados de todas as ações"""
        for ticker in self.acoes.keys():
            dados = self.buscar_dados_acao(ticker)
            if dados:
                self.dados.append(dados)

    def gerar_relatorio_excel(self, nome_arquivo='relatorio_investimentos.xlsx'):
        """Gera relatório em Excel"""
        if not self.dados:
            print("Nenhum dado disponível para gerar relatório")
            return

        df = pd.DataFrame(self.dados)

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Análise de Ações'

        # Estilos
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Adicionar header
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Adicionar dados
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Formatar números
                if col_idx > 2:  # Colunas numéricas
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.00'

        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

        # Salvar arquivo
        wb.save(nome_arquivo)
        print(f"\n✅ Relatório gerado com sucesso: {nome_arquivo}")
        print(f"   Localização: {os.path.abspath(nome_arquivo)}")

    def exibir_resumo(self):
        """Exibe um resumo dos dados coletados"""
        if not self.dados:
            print("Nenhum dado disponível")
            return

        df = pd.DataFrame(self.dados)
        print("\n" + "="*100)
        print("RELATÓRIO DE ANÁLISE DE AÇÕES".center(100))
        print("="*100)
        print(df.to_string(index=False))
        print("="*100)

    def executar(self, gerar_excel=True):
        """Executa o fluxo completo"""
        print("\n🚀 Iniciando análise de investimentos...\n")
        self.coletar_dados()
        self.exibir_resumo()

        if gerar_excel:
            self.gerar_relatorio_excel()


if __name__ == '__main__':
    assistente = InvestmentAssistant()
    assistente.executar()
